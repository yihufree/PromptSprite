# -*- coding: utf-8 -*-
"""
excel_io.py - Excel(.xlsx) 数据导入导出（批量编辑场景）
创建日期：2026-08-12（阶段六创建；阶段八重构为全局分类+领域关联 v2）

导出列：领域、一级分类、二级分类、名称、介绍、溯源、核心特征、应用场景、
        代表作、配图描述、提示词中文、提示词英文、图像获取方案、收藏
导入：按 领域/一级/二级 名称解析或新建全局分类，并建立 领域↔一级分类 关联（多对一共享）。
"""
from openpyxl import Workbook, load_workbook

from .. import config  # 2026-08-29（B3 修复）：新建根目录兜底归入"未明确分类"
from ..database import Database  # 2026-08-18（P1-1）：内容判重键 content_key
from ..models import Entry
from .json_io import _gather_categories, _chain_names

_HEADERS = ["领域", "一级分类", "二级分类", "名称", "介绍", "溯源", "核心特征",
            "应用场景", "代表作", "配图描述", "提示词中文", "提示词英文", "图像获取方案", "收藏"]


def resolve_category_path(db, domain_name: str, l1_name: str, l2_name: str):
    """按 领域/一级/二级 名称解析目标分类 id（多对一共享）：
    - 领域不存在则创建；一级分类已存在则仅建立领域关联，不存在则新建并关联；
    - 二级分类不存在则在一级下新建；全空返回 None(未分类)。
    """
    if not domain_name:
        return None
    domain = next((d for d in db.list_domains() if d["name"] == domain_name), None)
    if domain:
        did = domain["id"]
    else:
        # 2026-08-29（B3 修复）：新建根目录默认归入"未明确分类"，避免落"未分配"
        fallback = db.ensure_project(config.PROJECT_FALLBACK)
        did = db.add_domain(domain_name, project_id=fallback)
    if not l1_name:
        return None
    l1 = next((c for c in db.list_categories(parent_id=None) if c["name"] == l1_name), None)
    if l1 is None:
        l1_id = db.add_category(l1_name, domain_id=did)
    else:
        l1_id = l1["id"]
        db.link_domain_category(did, l1_id)  # 共享：已有的一级分类关联到该领域
    if not l2_name:
        return l1_id
    l2 = next((c for c in db.list_categories(parent_id=l1_id) if c["name"] == l2_name), None)
    return l2["id"] if l2 else db.add_category(l2_name, parent_id=l1_id)


# ---------------------------------------------------------------------- #
# 导出
# ---------------------------------------------------------------------- #
def _entry_domain_names(db, entry) -> str:
    """条目所属一级分类关联的领域名（多个用 / 分隔）"""
    root = db.category_root(entry["category_id"]) if entry["category_id"] else None
    if root is None:
        return ""
    return "/".join(d["name"] for d in db.linked_domains(root))


def export_excel(db, path, category_id=None) -> int:
    """导出全部（或指定分类子树）为 xlsx；返回导出的条目数"""
    if category_id is None:
        entries = db.list_all_entries()
    else:
        cat_ids = [c["id"] for c in _gather_categories(db, parent_id=category_id)]
        entries = [e for cid in cat_ids for e in db.list_entries(cid)]

    wb = Workbook()
    ws = wb.active
    ws.title = "提示词"
    ws.append(_HEADERS)
    for e in entries:
        chain = _chain_names(db, e["category_id"]) if e["category_id"] else []
        ws.append([
            _entry_domain_names(db, e),
            chain[0] if len(chain) > 0 else "",
            chain[1] if len(chain) > 1 else "",
            e["name"], e["intro"], e["origin"], e["features"], e["scenes"], e["works"],
            e["image_desc"], e["prompt_cn"], e["prompt_en"], e["image_plan"], e["is_favorite"],
        ])
    wb.save(path)
    return len(entries)


# ---------------------------------------------------------------------- #
# 导入
# ---------------------------------------------------------------------- #
def count_excel_rows(path) -> int:
    """预览：Excel 中除表头外的数据行数（用于进度条总量）"""
    wb = load_workbook(path, read_only=True)
    try:
        ws = wb.active
        return max((ws.max_row or 1) - 1, 0)
    finally:
        wb.close()


def import_excel(db, path, progress_cb=None) -> dict:
    """导入 xlsx；返回统计 {'entries': n}

    2026-08-29（P2 修复）：改用 read_only 流式读取，避免大文件内存暴涨。
    """
    wb = load_workbook(path, read_only=True)
    try:
        ws = wb.active
        header = [str(c.value).strip() if c.value else "" for c in ws[1]]
        col = {h: header.index(h) for h in _HEADERS if h in header}
        if "名称" not in col:
            raise ValueError("Excel 缺少【名称】列，请使用本软件导出的模板格式")

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        total = max(len(rows), 1)
        done, skipped, processed = 0, 0, 0
        pending = []
        seen_by_cat = {}  # category_id -> 该分类现有条目的"详情内容"键集合（P1-1 去重用）
        for r in rows:
            name = str(r[col["名称"]] or "").strip()
            if not name:
                continue
            cid = resolve_category_path(
                db,
                str(r[col["领域"]] or "").strip(),
                str(r[col["一级分类"]] or "").strip(),
                str(r[col["二级分类"]] or "").strip(),
            )
            e = Entry(category_id=cid, name=name,
                      intro=_cell(r, col, "介绍"), origin=_cell(r, col, "溯源"),
                      features=_cell(r, col, "核心特征"), scenes=_cell(r, col, "应用场景"),
                      works=_cell(r, col, "代表作"), image_desc=_cell(r, col, "配图描述"),
                      prompt_cn=_cell(r, col, "提示词中文"), prompt_en=_cell(r, col, "提示词英文"),
                      image_plan=_cell(r, col, "图像获取方案"),
                      is_favorite=_to_int(_cell(r, col, "收藏")))
            # 2026-08-18（P1-1）：详情内容去重——仅当内容完全相同才跳过，名称相同但内容不同仍新增
            keys = seen_by_cat.setdefault(cid, set())
            if not keys:
                keys.update(Database.content_key(x) for x in db.list_entries(cid))
            key = Database.content_key(e)
            processed += 1
            if key in keys:
                skipped += 1
                if progress_cb:
                    progress_cb(processed, total, f"{name}（重复跳过）")
                continue
            keys.add(key)
            pending.append(e)
            done += 1
            if progress_cb:
                progress_cb(processed, total, name)
        # 2026-08-18（P2-4）：改为收集后批量插入，单事务提交
        if pending:
            db.add_entries_batch(pending)
        return {"entries": done, "skipped": skipped}
    finally:
        wb.close()


def _cell(row, col_map, key) -> str:
    """安全读取单元格文本（不存在则返回空串）"""
    idx = col_map.get(key)
    if idx is None or idx >= len(row):
        return ""
    v = row[idx]
    return "" if v is None else str(v).strip()


def _to_int(v: str) -> int:
    """收藏列容错：兼容 1 / 1.0 / '1' / 是/√ 等写法，返回 0/1"""
    if not v:
        return 0
    try:
        return 1 if float(v) else 0
    except (ValueError, TypeError):
        return 1 if v.lower() in ("是", "true", "yes", "√", "✓") else 0
